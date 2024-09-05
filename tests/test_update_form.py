# Copyright (c) 2022, 2023 Humanitarian OpenStreetMap Team
#
# This file is part of osm_fieldwork.
#
#     osm-fieldwork is free software: you can redistribute it and/or modify
#     it under the terms of the GNU General Public License as published by
#     the Free Software Foundation, either version 3 of the License, or
#     (at your option) any later version.
#
#     osm-fieldwork is distributed in the hope that it will be useful,
#     but WITHOUT ANY WARRANTY; without even the implied warranty of
#     MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#     GNU General Public License for more details.
#
#     You should have received a copy of the GNU General Public License
#     along with osm_fieldwork.  If not, see <https:#www.gnu.org/licenses/>.
#
"""Test functionalty of update_form.py."""

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from osm_fieldwork.update_form import update_xls_form


def test_merge_mandatory_fields():
    """Merge the mandatory fields XLSForm to a test survey form."""
    test_form = Path(__file__).parent / "testdata" / "test_form_for_mandatory_fields.xls"
    with open(test_form, "rb") as xlsform:
        form_bytes = BytesIO(xlsform.read())

    updated_form = update_xls_form(form_bytes)
    workbook = load_workbook(filename=BytesIO(updated_form.getvalue()))

    # Check the 'survey' sheet
    if "survey" not in workbook.sheetnames:
        raise ValueError("The 'survey' sheet was not found in the workbook")
    survey_sheet = workbook["survey"]

    # Find the index of the 'name' column
    name_col = None
    for col in survey_sheet.iter_cols(1, survey_sheet.max_column):
        if col[0].value == "name":
            name_col = col
            break
    assert name_col is not None, "The 'name' column was not found."

    # Check if certain fields are present in the 'name' column (skip the header)
    existing_field = any(cell.value == "existing" for cell in name_col[1:])
    assert existing_field, "'existing' field not found in the 'name' column."

    status_field = any(cell.value == "status" for cell in name_col[1:])
    assert status_field, "'status' field not found in the 'name' column."

    digitisation_correct_field = any(cell.value == "digitisation_correct" for cell in name_col[1:])
    assert digitisation_correct_field, "'digitisation_correct' field not found in the 'name' column."

    # Check that the 'name' column does not have a duplicate entry for 'username'
    username_count = sum(1 for cell in name_col[1:] if cell.value == "username")
    assert username_count <= 1, "Duplicate 'username' entry found in the 'name' column."

    # Check the 'choices' sheet
    if "choices" not in workbook.sheetnames:
        raise ValueError("The 'choices' sheet was not found in the workbook")
    choices_sheet = workbook["choices"]

    # Find the index of the 'name' column in the 'choices' sheet
    choices_name_col = None
    for col in choices_sheet.iter_cols(1, choices_sheet.max_column):
        if col[0].value == "name":
            choices_name_col = col
            break

    assert choices_name_col is not None, "'name' column was not found in the 'choices' sheet."

    # Test: Check that the 'choices' sheet does not have duplicate entries for 'yes' and 'no'
    yes_count = sum(1 for cell in choices_name_col[1:] if cell.value == "yes")
    no_count = sum(1 for cell in choices_name_col[1:] if cell.value == "no")
    assert yes_count <= 1, "Duplicate 'yes' entry found in the 'value' column of 'choices' sheet."
    assert no_count <= 1, "Duplicate 'no' entry found in the 'value' column of 'choices' sheet."

    # Check the 'entities' sheet
    if "entities" not in workbook.sheetnames:
        raise ValueError("The 'entities' sheet was not found in the workbook")
    entities_sheet = workbook["entities"]

    # Find the index of the 'label' column in the 'entities' sheet
    entities_label_col = None
    for col in entities_sheet.iter_cols(1, entities_sheet.max_column):
        if col[0].value == "label":
            entities_label_col = col
            break

    assert entities_label_col is not None, "'label' column was not found in the 'entities' sheet."

    # Check that the 'entities' label value of 'test label' is replaced by required value
    test_label_present = any(cell.value == "test label" for cell in entities_label_col[1:])
    assert not test_label_present, "'test label' found in the 'label' column of 'entities' sheet."

    # TODO add test to check that digitisation questions come at end of sheet
